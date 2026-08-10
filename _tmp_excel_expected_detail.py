# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook

base = Path(r"C:\Users\testii\Downloads\dash\DashboardBack\temp")
target = None
for p in base.glob("*.xlsx"):
    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    text = " ".join(str(c) for c in first if c)
    wb.close()
    if "01.08.2026" in text:
        target = p
        break

wb = load_workbook(target, data_only=True)
ws = wb.active

# Find dept row ranges
depts = []
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if isinstance(a, str) and (a.startswith("Отдел") or a.startswith("(ликв")):
        depts.append((r, a, ws.cell(r, 4).value, ws.cell(r, 6).value, ws.cell(r, 8).value))
print("DEPTS:")
for d in depts:
    print(d)

# For each dept with expected, print detail rows where col6 has value
depts.append((ws.max_row + 1, "END", None, None, None))
for i in range(len(depts) - 1):
    r0, name, plan, exp, fact = depts[i]
    r1 = depts[i + 1][0]
    if not isinstance(exp, (int, float)) or not exp:
        continue
    print(f"\n=== {name} expected={exp} detail col6 ===")
    s = 0
    for r in range(r0 + 1, r1):
        v = ws.cell(r, 6).value
        if isinstance(v, (int, float)) and v:
            print(r, ws.cell(r, 1).value, v, "fact8", ws.cell(r, 8).value)
            s += float(v)
    print("detail sum", s)
