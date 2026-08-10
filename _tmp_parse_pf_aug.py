# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook

base = Path(r"C:\Users\testii\Downloads\dash\DashboardBack\temp")
# pick the august file by period text
target = None
for p in base.glob("*.xlsx"):
    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception:
        continue
    ws = wb[wb.sheetnames[0]]
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    text = " ".join(str(c) for c in first if c)
    wb.close()
    if "01.08.2026" in text or "01.08.26" in text:
        target = p
        print("TARGET", p.name, text)
        break
if not target:
    raise SystemExit("no aug file")

wb = load_workbook(target, data_only=True)
ws = wb[wb.sheetnames[0]]

# print header row 8 with column letters
print("HEADER row8:")
for col in range(1, 30):
    v = ws.cell(8, col).value
    if v is not None:
        print(f"  col{col}: {v}")

print("\nHEADER row7:")
for col in range(1, 30):
    v = ws.cell(7, col).value
    if v is not None:
        print(f"  col{col}: {v}")

# Find department summary rows: cells in col1 containing 'Отдел'
print("\nDEPT ROWS:")
totals = {}
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if not a or not isinstance(a, str):
        continue
    if a.startswith("Отдел") or a.startswith("(ликв"):
        vals = [ws.cell(r, c).value for c in range(1, 20)]
        print(r, vals)

# Try to find grand total row
print("\nPOSSIBLE TOTALS:")
for r in range(1, min(ws.max_row, 400) + 1):
    a = ws.cell(r, 1).value
    if a and isinstance(a, str) and ("Итог" in a or "Всего" in a or a.strip() == ""):
        vals = [ws.cell(r, c).value for c in range(1, 16)]
        if any(isinstance(v, (int, float)) for v in vals):
            print(r, vals)
