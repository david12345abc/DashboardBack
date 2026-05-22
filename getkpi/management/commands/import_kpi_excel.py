"""
Импорт KPI из Excel-методики (лист с колонками «Блок дашборда», «Код элемента», …).

Использование:
    py manage.py import_kpi_excel --file "../Лист Microsoft Excel.xlsx" --department "Начальник отдела автоматизации ИТ"
    py manage.py import_kpi_excel --file path.xlsx --department "…" --code-prefix ИТ --dry-run
"""
from __future__ import annotations

import re
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from getkpi.models import KpiDefinition

DEFAULT_FILE = Path(__file__).resolve().parents[4] / "Лист Microsoft Excel.xlsx"

COL = {
    "block": 0,
    "position": 1,
    "kpi_id": 2,
    "frequency": 3,
    "name": 4,
    "element_type": 5,
    "goal": 6,
    "formula": 7,
    "unit": 8,
    "source": 9,
    "green": 13,
    "yellow": 14,
    "red": 15,
}

BLOCK_MAP = {
    "Плитки KPI": "плитка",
    "Графики": "график",
    "Таблицы": "таблица",
}

CHART_TYPE_BY_SUFFIX = {
    "C1": ("multi_line_plan_fact_monthly", "График тренда"),
    "C2": ("column_plan_fact_waterfall_quarterly", "План/факт / waterfall"),
    "C3": ("heatmap_rag", "Heatmap / структура"),
}

FIELDS = (
    "name", "block", "frequency", "perspective", "goal",
    "formula", "unit", "source",
    "monthly_target", "quarterly_target", "yearly_target",
    "green_threshold", "yellow_threshold", "red_threshold",
    "weight_pct", "chart_type", "chart_type_label",
)


def _cell(row: tuple, idx: int):
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    text = str(val).strip()
    return text or None


def _norm_freq(freq: str | None) -> str:
    if not freq:
        return ""
    low = freq.lower()
    if "ежемесяч" in low:
        return "Ежемесячно"
    if "ежеквартал" in low:
        return "Ежеквартально"
    if "ежегод" in low:
        return "Ежегодно"
    return freq


def _infer_perspective(block: str, element_type: str | None, name: str | None) -> str:
    blob = " ".join(x for x in (element_type, name) if x).lower()
    if block == "график" or block == "таблица":
        return "Управление"
    if any(x in blob for x in ("бюджет", "фот", "рентабель", "финанс")):
        return "Финансы"
    if any(x in blob for x in ("текучест", "персонал")):
        return "Персонал"
    if any(x in blob for x in ("csi", "nps", "лояльност", "цифров", "автоматизац", "инициатив")):
        return "Развитие"
    if any(x in blob for x in ("sla", "заявк", "доступност", "систем")):
        return "Процессы"
    if any(x in blob for x in ("клиент", "сервис")):
        return "Клиенты"
    return "Процессы"


def _period_targets(frequency: str, green: str | None) -> dict[str, str | None]:
    if not green:
        return {"monthly_target": None, "quarterly_target": None, "yearly_target": None}
    low = frequency.lower()
    if "ежемесяч" in low:
        return {"monthly_target": green, "quarterly_target": None, "yearly_target": green}
    if "ежеквартал" in low:
        return {"monthly_target": None, "quarterly_target": green, "yearly_target": green}
    if "ежегод" in low:
        return {"monthly_target": None, "quarterly_target": None, "yearly_target": green}
    return {"monthly_target": green, "quarterly_target": None, "yearly_target": None}


def _replace_code_prefix(kpi_id: str, *, src_prefix: str, dst_prefix: str) -> str:
    if kpi_id.startswith(f"{src_prefix}-"):
        return f"{dst_prefix}-{kpi_id[len(src_prefix) + 1:]}"
    return kpi_id


def _detect_src_prefix(kpi_ids: list[str]) -> str | None:
    for kpi_id in kpi_ids:
        m = re.match(r"^([A-ZА-ЯЁ]{2,5})-", kpi_id)
        if m:
            return m.group(1)
    return None


def _parse_excel(path: Path, *, code_prefix: str | None, src_prefix: str | None) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [_cell(rows[0], i) for i in range(len(rows[0]))]
    if header[COL["block"]] != "Блок дашборда":
        raise ValueError("Не найден заголовок «Блок дашборда» — проверьте формат Excel.")

    raw_ids = [
        _cell(row, COL["kpi_id"])
        for row in rows[1:]
        if _cell(row, COL["kpi_id"])
    ]
    detected = _detect_src_prefix(raw_ids)
    src = src_prefix or detected or "SRV"
    dst = code_prefix or "ИТ"

    parsed: list[dict] = []
    for row in rows[1:]:
        block_raw = _cell(row, COL["block"])
        kpi_id_raw = _cell(row, COL["kpi_id"])
        if not block_raw or not kpi_id_raw:
            continue
        if block_raw == "Шапка":
            continue

        block = BLOCK_MAP.get(block_raw)
        if not block:
            continue

        kpi_id = _replace_code_prefix(kpi_id_raw, src_prefix=src, dst_prefix=dst)
        frequency = _norm_freq(_cell(row, COL["frequency"]))
        name = _cell(row, COL["name"]) or kpi_id
        goal = _cell(row, COL["goal"]) or ""
        formula = _cell(row, COL["formula"]) or ""
        unit = _cell(row, COL["unit"]) or ""
        source = _cell(row, COL["source"]) or ""
        green = _cell(row, COL["green"])
        yellow = _cell(row, COL["yellow"])
        red = _cell(row, COL["red"])
        element_type = _cell(row, COL["element_type"])

        item = {
            "kpi_id": kpi_id,
            "name": name,
            "block": block,
            "frequency": frequency,
            "perspective": _infer_perspective(block, element_type, name),
            "goal": goal,
            "formula": formula,
            "unit": unit,
            "source": source,
            "green_threshold": green,
            "yellow_threshold": yellow,
            "red_threshold": red,
            "weight_pct": None,
            **_period_targets(frequency, green),
        }

        suffix = kpi_id.split("-")[-1] if "-" in kpi_id else ""
        if block == "график" and suffix in CHART_TYPE_BY_SUFFIX:
            chart_type, chart_label = CHART_TYPE_BY_SUFFIX[suffix]
            item["chart_type"] = chart_type
            item["chart_type_label"] = chart_label

        parsed.append(item)
    return parsed


class Command(BaseCommand):
    help = "Импорт KPI-определений из Excel-методики в kpi_definition"

    def add_arguments(self, parser):
        parser.add_argument("--file", type=str, default=str(DEFAULT_FILE), help="Путь к .xlsx")
        parser.add_argument("--department", type=str, required=True, help="Подразделение (department)")
        parser.add_argument("--code-prefix", type=str, default="ИТ", help="Префикс кодов KPI в БД")
        parser.add_argument("--src-prefix", type=str, default=None, help="Исходный префикс в Excel (авто)")
        parser.add_argument("--dry-run", action="store_true", help="Только показать изменения")

    def handle(self, *args, **options):
        path = Path(options["file"])
        department = options["department"].strip()
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"Файл не найден: {path}"))
            return

        try:
            kpi_list = _parse_excel(
                path,
                code_prefix=options["code_prefix"],
                src_prefix=options["src_prefix"],
            )
        except ValueError as exc:
            self.stderr.write(self.style.ERROR(str(exc)))
            return

        if not kpi_list:
            self.stderr.write(self.style.ERROR("В файле нет строк KPI для импорта."))
            return

        dry_run = bool(options["dry_run"])
        desired_ids = {item["kpi_id"] for item in kpi_list}
        existing = list(
            KpiDefinition.objects.filter(department=department).order_by("position", "kpi_id")
        )
        stale = [item for item in existing if item.kpi_id not in desired_ids]

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN"))
            self.stdout.write(f"Подразделение: {department}")
            self.stdout.write(f"Строк из Excel: {len(kpi_list)}")
            for item in stale:
                self.stdout.write(f"  DELETE {item.kpi_id} — {item.name}")
            for pos, item in enumerate(kpi_list):
                exists = KpiDefinition.objects.filter(
                    department=department, kpi_id=item["kpi_id"],
                ).exists()
                tag = "UPDATE" if exists else "CREATE"
                self.stdout.write(f"  [{tag}] pos={pos} {item['kpi_id']} — {item['name']}")
            return

        with transaction.atomic():
            if stale:
                KpiDefinition.objects.filter(id__in=[item.id for item in stale]).delete()

            created = updated = 0
            for pos, kpi in enumerate(kpi_list):
                defaults = {"position": pos}
                for field in FIELDS:
                    val = kpi.get(field)
                    if val is not None:
                        defaults[field] = val
                _, is_new = KpiDefinition.objects.update_or_create(
                    department=department,
                    kpi_id=kpi["kpi_id"],
                    defaults=defaults,
                )
                if is_new:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"Импорт завершён для «{department}»: {len(kpi_list)} KPI "
            f"({created} новых, {updated} обновлённых, {len(stale)} удалено)"
        ))
