"""
Таблица «группы / двойки» ОДП из файла выгрузки 1С (формат как «ОДП март.xlsx»).

В публикации OData у Document_ЗаказКлиента нет пары ФИО как в отчёте (Менеджер — пользователь,
Руководитель — ответственные лица орг., это не вторая «половина» двойки). Поэтому детализация
как в вашем Excel подключается из файлов в каталоге getkpi/temp.

Имя файла: «ОДП {имя месяца}.xlsx», месяц — русское название в нижнем регистре (январь … декабрь).

Итоговая строка листа: первый столбец ровно «Отдел дилерских продаж», факты в столбцах H, Q, Y.
Группы: строка начинается с «БЫТ» или «ПРОМ» и содержит разделитель « / » или неразрывный пробел+слэш.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

MONTH_FILE_NAMES = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}

# 1-based номера столбцов на типовом листе (см. шапку «Договоры заключенные (факт)» и т.д.)
COL_DOG_1BASE = 8
COL_MONEY_1BASE = 17
COL_SHIP_1BASE = 25

DEPT_TOTAL_LABEL = "Отдел дилерских продаж"

_GROUP_RE = re.compile(r"^(БЫТ|ПРОМ)\s+.+", re.DOTALL)


def _temp_dir() -> Path:
    return Path(__file__).resolve().parent / "temp"


def _workbook_path(year: int, month: int) -> Path:
    mname = MONTH_FILE_NAMES.get(month, str(month))
    return _temp_dir() / f"ОДП {mname}.xlsx"


def _is_manager_group_row(cell_a: str) -> bool:
    s = (cell_a or "").strip().replace("\r", "").replace("\n", " ")
    if not s:
        return False
    if not _GROUP_RE.match(s):
        return False
    return "\u00a0/\u00a0" in s or " / " in s


def load_odp_reference_table(year: int, month: int) -> dict[str, Any] | None:
    """Прочитать выгрузку ОДП из temp. Невозможно — None (нет файла / библиотеки / формата)."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("odp_excel_breakdown: openpyxl не установлен")
        return None

    path = _workbook_path(year, month)
    if not path.is_file():
        return None

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        logger.exception("odp_excel_breakdown: не удалось открыть %s", path)
        return None

    dept_totals: dict[str, float | None] = {"dogovory": None, "dengi": None, "otgruzki": None}
    groups: list[dict[str, Any]] = []

    try:
        for rno in range(1, ws.max_row + 1):
            c0 = ws.cell(rno, 1).value
            if c0 is None:
                continue
            label = str(c0).strip()
            if label == DEPT_TOTAL_LABEL:
                dept_totals["dogovory"] = _num(ws.cell(rno, COL_DOG_1BASE).value)
                dept_totals["dengi"] = _num(ws.cell(rno, COL_MONEY_1BASE).value)
                dept_totals["otgruzki"] = _num(ws.cell(rno, COL_SHIP_1BASE).value)
                continue
            if _is_manager_group_row(label):
                groups.append({
                    "group": _normalize_label(label),
                    "dogovory": _num(ws.cell(rno, COL_DOG_1BASE).value),
                    "dengi": _num(ws.cell(rno, COL_MONEY_1BASE).value),
                    "otgruzki": _num(ws.cell(rno, COL_SHIP_1BASE).value),
                })
    finally:
        wb.close()

    if dept_totals["dogovory"] is None and not groups:
        return None

    gsum_d = sum((g["dogovory"] or 0) for g in groups)
    gsum_m = sum((g["dengi"] or 0) for g in groups)
    gsum_s = sum((g["otgruzki"] or 0) for g in groups)
    dtot = float(dept_totals["dogovory"] or 0)
    note = None
    if dept_totals["dogovory"] is not None and groups and abs(gsum_d - dtot) > 0.05:
        note = (
            f"Сумма по строкам групп ({round(gsum_d, 2)}) не совпадает со строкой «{DEPT_TOTAL_LABEL}» "
            f"({round(dtot, 2)}) — как в исходном Excel."
        )

    return {
      "kpi_id": "ODP-T-REFERENCE",
      "name": "ОДП — группы (выгрузка 1С)",
      "description": (
          "Показатели и состав групп взяты из файла выгрузки в каталоге getkpi/temp. "
          "Плитки KPI считаются из OData и могут расходиться, пока правила отчёта полностью не совпадут."
      ),
      "period": {
          "year": year,
          "month": month,
          "month_name": MONTH_FILE_NAMES.get(month, str(month)),
      },
      "source_file": path.name,
      "totals_row": dept_totals,
      "columns": [
          "Группа (как в отчёте)",
          "Договоры (факт), руб.",
          "Деньги (факт), руб.",
          "Отгрузки (факт), руб.",
      ],
      "rows": [
          {
              "Группа (как в отчёте)": g["group"],
              "Договоры (факт), руб.": g["dogovory"],
              "Деньги (факт), руб.": g["dengi"],
              "Отгрузки (факт), руб.": g["otgruzki"],
          }
          for g in groups
      ],
      "groups_row_count": len(groups),
      "groups_sum": {"dogovory": round(gsum_d, 2), "dengi": round(gsum_m, 2), "otgruzki": round(gsum_s, 2)},
      "reconciliation_note": note,
    }


def _num(v: Any) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _normalize_label(s: str) -> str:
    s = s.strip()
    s = s.replace("\u00a0", " ")
    s = re.sub(r"\s+", " ", s)
    return s
