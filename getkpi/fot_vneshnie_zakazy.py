"""
План/факт внешних заказов из fot_vneshnie_zakazy.xlsx (строка «Итого» по месяцам).

Используется TD-M6 и сводными выгрузками (плитка TD-M5 — TurboProject).
"""
from __future__ import annotations

import logging
import os
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

DEFAULT_XLSX_PATH = Path(__file__).resolve().parents[2] / "fot_vneshnie_zakazy.xlsx"

_MONTH_HEADER_STEMS: tuple[tuple[str, int], ...] = (
    ("январ", 1),
    ("феврал", 2),
    ("март", 3),
    ("апрел", 4),
    ("май", 5),
    ("июн", 6),
    ("июл", 7),
    ("август", 8),
    ("сентябр", 9),
    ("октябр", 10),
    ("ноябр", 11),
    ("декабр", 12),
)

_TOTAL_ROW_LABELS = frozenset({"итого"})


def resolve_xlsx_path(explicit: Path | str | None = None) -> Path:
    if explicit is not None:
        return Path(explicit)
    env_path = os.getenv("FOT_VNESHNIE_ZAKAZY_XLSX")
    if env_path:
        return Path(env_path)
    return DEFAULT_XLSX_PATH


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    return num if num == num else None


def _month_from_header_cell(value: Any, year: int) -> int | None:
    text = str(value or "").lower().replace("ё", "е")
    if str(year) not in text:
        return None
    for stem, month in _MONTH_HEADER_STEMS:
        if stem in text:
            return month
    return None


def _parse_month_columns(ws: Any, year: int) -> dict[int, tuple[int, int]]:
    month_cols: dict[int, tuple[int, int]] = {}
    for col in range(1, ws.max_column + 1):
        month = _month_from_header_cell(ws.cell(1, col).value, year)
        if month is not None and month not in month_cols:
            month_cols[month] = (col, col + 1)
    return month_cols


def _find_total_row_index(ws: Any) -> int | None:
    for row_idx in range(1, ws.max_row + 1):
        label_raw = ws.cell(row_idx, 1).value
        if label_raw is None:
            continue
        if str(label_raw).strip().lower().replace("ё", "е") in _TOTAL_ROW_LABELS:
            return row_idx
    return None


@lru_cache(maxsize=8)
def _load_workbook_cached(path_str: str, mtime_ns: int) -> Any:
    return load_workbook(path_str, data_only=True)


def load_monthly_totals_row(
    year: int,
    *,
    path: Path | str | None = None,
) -> dict[int, dict[str, float | None]]:
    """
    Помесячные итоги из строки «Итого»:
    ``{month: {"plan": ..., "fact": ...}}``.
    """
    xlsx_path = resolve_xlsx_path(path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Не найден файл: {xlsx_path}")

    stat = xlsx_path.stat()
    wb = _load_workbook_cached(str(xlsx_path.resolve()), stat.st_mtime_ns)
    ws = wb.active
    month_cols = _parse_month_columns(ws, year)
    if not month_cols:
        raise ValueError(f"В {xlsx_path} нет заголовков месяцев за {year} г.")

    total_row = _find_total_row_index(ws)
    if total_row is None:
        raise ValueError(f"В {xlsx_path} не найдена строка «Итого».")

    totals: dict[int, dict[str, float | None]] = {}
    for month, (plan_col, fact_col) in month_cols.items():
        totals[month] = {
            "plan": _safe_float(ws.cell(total_row, plan_col).value),
            "fact": _safe_float(ws.cell(total_row, fact_col).value),
        }
    return totals


def load_project_rows(
    year: int,
    *,
    path: Path | str | None = None,
) -> list[dict[str, Any]]:
    """Строки проектов (без «Итого») с plan/fact по месяцам — для сводных выгрузок."""
    xlsx_path = resolve_xlsx_path(path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Не найден файл: {xlsx_path}")

    stat = xlsx_path.stat()
    wb = _load_workbook_cached(str(xlsx_path.resolve()), stat.st_mtime_ns)
    ws = wb.active
    month_cols = _parse_month_columns(ws, year)

    skip_labels = _TOTAL_ROW_LABELS | frozenset({
        "проект",
        "статья бюджета",
        "регистратор",
    })
    rows: list[dict[str, Any]] = []
    for row_idx in range(3, ws.max_row + 1):
        label_raw = ws.cell(row_idx, 1).value
        if label_raw is None or str(label_raw).strip() == "":
            continue
        label = str(label_raw).strip()
        if label.lower().replace("ё", "е") in skip_labels:
            continue
        months: dict[int, dict[str, float | None]] = {}
        for month, (plan_col, fact_col) in month_cols.items():
            months[month] = {
                "plan": _safe_float(ws.cell(row_idx, plan_col).value),
                "fact": _safe_float(ws.cell(row_idx, fact_col).value),
            }
        rows.append({"label": label, "months": months})
    return rows
