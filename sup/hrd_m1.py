"""HRD-M1 — критические вакансии типа A, закрытые в срок.

Источник: ``sup/SUP_data.xlsx``, лист ``Вакансии``.
План месяца — закрытые фактом вакансии типа A без исключений.
Факт месяца — часть плана, закрытая не позже плановой даты.
"""
from __future__ import annotations

import json
import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from getkpi.cache_manager import locked_call
from getkpi.devdir import ytd_json_cache
from getkpi.devdir.rd_monthly_period import MONTH_NAMES, normalize_rd_tile_period

logger = logging.getLogger(__name__)

SOURCE_FILE = Path(__file__).resolve().parent / "SUP_data.xlsx"
SHEET_NAME = "Вакансии"
CACHE_PREFIX = "sup_hrd_m1_vacancies"
CACHE_SOURCE_TAG = "sup_hrd_m1_vacancies_payload_v1"
CACHE_VERSION = 2

EXCLUSION_PHRASES = (
    "снята заказчиком",
    "снят заказчиком",
    "снято заказчиком",
    "отменена",
    "отменен",
    "отменено",
    "приостановлена",
    "приостановлен",
    "приостановлено",
    "заморожена",
    "заморожен",
    "заморожено",
    "закрыта заказчиком",
    "закрыт заказчиком",
    "закрыто заказчиком",
    "дубль",
    "дубликат",
    "ошибка заявки",
)


def _normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return " ".join(re.sub(r"[^0-9a-zа-я]+", " ", text).split())


def _normalize_header(value: Any) -> str:
    return _normalize_text(value).replace(" ", "")


def _normalize_type(value: Any) -> str:
    return str(value or "").strip().upper().replace("А", "A")


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 20_000:
        try:
            return from_excel(value).date()
        except (TypeError, ValueError):
            return None
    raw = str(value).strip()
    if not raw:
        return None
    for candidate in (raw, raw[:10]):
        try:
            return datetime.fromisoformat(candidate.replace("Z", "+00:00")).date()
        except ValueError:
            pass
        if len(candidate) >= 10 and candidate[2:3] == "." and candidate[5:6] == ".":
            try:
                return datetime.strptime(candidate[:10], "%d.%m.%Y").date()
            except ValueError:
                pass
    return None


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        if not value:
            return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    return num if num == num else None


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    required = {"авс", "датазакрытияфакт"}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = {
            _normalize_header(value): idx
            for idx, value in enumerate(row)
            if value is not None and str(value).strip()
        }
        if required.issubset(headers):
            return row_idx, headers
    raise RuntimeError("Не найдена строка заголовков листа 'Вакансии'")


def _row_value(row: tuple[Any, ...], headers: dict[str, int], header_key: str) -> Any:
    idx = headers.get(header_key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _exclusion_reason(row: tuple[Any, ...], headers: dict[str, int]) -> str | None:
    text_parts: list[str] = []
    for key in ("кандидатфио", "комментарии"):
        value = _row_value(row, headers, key)
        if value is not None:
            text_parts.append(str(value))
    for header, idx in headers.items():
        if "статус" in header and idx < len(row) and row[idx] is not None:
            text_parts.append(str(row[idx]))
    haystack = _normalize_text(" ".join(text_parts))
    for phrase in EXCLUSION_PHRASES:
        if _normalize_text(phrase) in haystack:
            return phrase
    return None


def _closed_on_time(
    fact_date: date,
    plan_date: date | None,
    deviation_value: Any,
) -> bool:
    if plan_date is not None:
        return fact_date <= plan_date
    deviation = _safe_float(deviation_value)
    return deviation is not None and deviation >= 0


def _load_vacancy_rows() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    header_row, headers = _find_header_row(ws)

    included: list[dict[str, Any]] = []
    excluded_by_month: dict[int, int] = {m: 0 for m in range(1, 13)}
    excluded_reasons: dict[str, int] = {}
    skipped_no_fact = 0
    skipped_non_a = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vacancy_type = _normalize_type(_row_value(row, headers, "авс"))
        if vacancy_type != "A":
            skipped_non_a += 1
            continue

        fact_date = _parse_date(_row_value(row, headers, "датазакрытияфакт"))
        if fact_date is None:
            skipped_no_fact += 1
            continue

        reason = _exclusion_reason(row, headers)
        if reason:
            excluded_by_month[fact_date.month] += 1
            excluded_reasons[reason] = excluded_reasons.get(reason, 0) + 1
            continue

        plan_date = _parse_date(_row_value(row, headers, "датазакрытияплановая"))
        included.append({
            "fact_date": fact_date,
            "plan_date": plan_date,
            "on_time": _closed_on_time(
                fact_date,
                plan_date,
                _row_value(row, headers, "отклоненияотсрока"),
            ),
        })

    return included, {
        "source_file": str(SOURCE_FILE),
        "sheet": SHEET_NAME,
        "header_row": header_row,
        "included_total": len(included),
        "skipped_non_a": skipped_non_a,
        "skipped_no_fact": skipped_no_fact,
        "excluded_total": sum(excluded_by_month.values()),
        "excluded_by_month": excluded_by_month,
        "excluded_reasons": excluded_reasons,
    }


def _build_payload(year: int | None = None, month: int | None = None) -> dict[str, Any]:
    ref_y, ref_m = normalize_rd_tile_period(year, month)
    vacancies, debug = _load_vacancy_rows()

    monthly_rows: list[dict[str, Any]] = []
    for m in range(1, ref_m + 1):
        month_items = [
            item
            for item in vacancies
            if item["fact_date"].year == ref_y and item["fact_date"].month == m
        ]
        plan = len(month_items)
        fact = sum(1 for item in month_items if item["on_time"])
        monthly_rows.append({
            "month": m,
            "year": ref_y,
            "month_name": MONTH_NAMES[m],
            "plan": plan,
            "fact": fact,
            "kpi_pct": round(fact / plan * 100, 2) if plan > 0 else None,
            "has_data": plan > 0,
            "values_unit": "шт.",
        })

    ref_row = monthly_rows[-1] if monthly_rows else None
    return {
        "data_granularity": "monthly",
        "monthly_data": monthly_rows,
        "last_full_month_row": dict(ref_row) if ref_row else None,
        "kpi_period": {
            "type": "last_full_month",
            "year": ref_y,
            "month": ref_m,
            "month_name": MONTH_NAMES[ref_m],
        },
        "ytd": {
            "total_plan": ref_row.get("plan") if ref_row else None,
            "total_fact": ref_row.get("fact") if ref_row else None,
            "kpi_pct": ref_row.get("kpi_pct") if ref_row else None,
            "months_with_data": sum(1 for row in monthly_rows if row.get("has_data")),
            "months_total": len(monthly_rows),
            "values_unit": "шт.",
        },
        "reference_analytics": {
            "excluded_total": debug["excluded_total"],
            "excluded_by_month": {
                m: debug["excluded_by_month"].get(m, 0)
                for m in range(1, ref_m + 1)
            },
            "excluded_reasons": debug["excluded_reasons"],
            "note": "Исключённые вакансии не входят в план/факт HRD-M1.",
        },
        "debug": {
            "kpi_id": "HRD-M1",
            "status": "ok",
            "rule": "type A vacancies closed by fact date; fact = closed on or before plan date",
            **debug,
        },
    }


def cache_file_path_for_period(year: int | None = None, month: int | None = None) -> Path:
    return ytd_json_cache.public_cache_path(CACHE_PREFIX, year, month)


def _load_cache(path: Path, *, source_mtime_ns: int, perpetual: bool) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            raw = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    if raw.get("cache_source") != CACHE_SOURCE_TAG:
        return None
    if raw.get("cache_version") != CACHE_VERSION:
        return None
    if raw.get("source_mtime_ns") != source_mtime_ns:
        return None
    payload = raw.get("payload")
    if not isinstance(payload, dict):
        return None
    if perpetual or raw.get("cache_date") == date.today().isoformat():
        return payload
    return None


def _save_cache(path: Path, payload: dict[str, Any], *, source_mtime_ns: int) -> None:
    ytd_json_cache.CACHE_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", encoding="utf-8") as f:
            json.dump(
                {
                    "cache_source": CACHE_SOURCE_TAG,
                    "cache_version": CACHE_VERSION,
                    "cache_date": date.today().isoformat(),
                    "source_mtime_ns": source_mtime_ns,
                    "payload": payload,
                },
                f,
                ensure_ascii=False,
                indent=2,
            )
    except OSError:
        logger.exception("HRD-M1: не удалось сохранить кэш")


def get_hrd_m1_ytd(year: int | None = None, month: int | None = None) -> dict[str, Any] | None:
    ref_y, ref_m = normalize_rd_tile_period(year, month)
    cache_path = ytd_json_cache.cache_path(CACHE_PREFIX, ref_y, ref_m)
    perpetual = ytd_json_cache.is_ref_period_fully_past(ref_y, ref_m)

    def _runner() -> dict[str, Any] | None:
        try:
            source_mtime_ns = SOURCE_FILE.stat().st_mtime_ns
        except OSError:
            logger.exception("HRD-M1: не найден источник %s", SOURCE_FILE)
            return None

        cached = _load_cache(cache_path, source_mtime_ns=source_mtime_ns, perpetual=perpetual)
        if cached is not None:
            return cached

        try:
            payload = _build_payload(year=ref_y, month=ref_m)
        except Exception:
            logger.exception("HRD-M1: ошибка расчёта вакансий")
            return None
        _save_cache(cache_path, payload, source_mtime_ns=source_mtime_ns)
        return payload

    return locked_call(f"sup_hrd_m1_{ref_y}_{ref_m:02d}", _runner)
