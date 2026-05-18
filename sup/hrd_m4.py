"""HRD-M4 — текучесть персонала по компании.

Источник: ``sup/SUP_data.xlsx``, лист ``Текучесть``.
Факт — строка 10, план — строка 11. Месяцы идут с января по колонкам H, K, N...
"""
from __future__ import annotations

import json
import logging
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from getkpi.cache_manager import locked_call
from getkpi.devdir import ytd_json_cache
from getkpi.devdir.rd_monthly_period import MONTH_NAMES, normalize_rd_tile_period

logger = logging.getLogger(__name__)

SOURCE_FILE = Path(__file__).resolve().parent / "SUP_data.xlsx"
SHEET_NAME = "Текучесть"
CACHE_PREFIX = "sup_hrd_m4_turnover"
CACHE_SOURCE_TAG = "sup_hrd_m4_turnover_payload_v2"
CACHE_VERSION = 2

FACT_ROW = 10
PLAN_ROW = 11
FIRST_MONTH_COLUMN = 8  # H
MONTH_COLUMN_STEP = 3


def _month_column(month: int) -> int:
    return FIRST_MONTH_COLUMN + (month - 1) * MONTH_COLUMN_STEP


def _safe_percent(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        raw = value.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
        if not raw or raw.startswith("#"):
            return None
        has_percent_sign = raw.endswith("%")
        if has_percent_sign:
            raw = raw[:-1]
        try:
            num = float(raw)
        except ValueError:
            return None
        if has_percent_sign:
            return round(num, 2)
    else:
        try:
            num = float(value)
        except (TypeError, ValueError):
            return None

    if num != num:
        return None
    if abs(num) <= 1:
        num *= 100
    return round(num, 2)


def _load_turnover_months(ref_y: int, ref_m: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    monthly_rows: list[dict[str, Any]] = []
    raw_cells: list[dict[str, Any]] = []

    for month in range(1, ref_m + 1):
        column = _month_column(month)
        fact_cell = ws.cell(FACT_ROW, column)
        plan_cell = ws.cell(PLAN_ROW, column)
        fact = _safe_percent(fact_cell.value)
        plan = _safe_percent(plan_cell.value)

        monthly_rows.append({
            "month": month,
            "year": ref_y,
            "month_name": MONTH_NAMES[month],
            "plan": plan,
            "fact": fact,
            "kpi_pct": fact,
            "has_data": fact is not None or plan is not None,
            "values_unit": "%",
        })
        raw_cells.append({
            "month": month,
            "fact_cell": fact_cell.coordinate,
            "plan_cell": plan_cell.coordinate,
            "raw_fact": fact_cell.value,
            "raw_plan": plan_cell.value,
            "fact": fact,
            "plan": plan,
        })

    return monthly_rows, {
        "source_file": str(SOURCE_FILE),
        "sheet": SHEET_NAME,
        "fact_row": FACT_ROW,
        "plan_row": PLAN_ROW,
        "first_month_column": "H",
        "month_column_step": MONTH_COLUMN_STEP,
        "raw_cells": raw_cells,
    }


def _build_payload(year: int | None = None, month: int | None = None) -> dict[str, Any]:
    ref_y, ref_m = normalize_rd_tile_period(year, month)
    monthly_rows, debug = _load_turnover_months(ref_y, ref_m)
    ref_row = monthly_rows[-1] if monthly_rows else None

    return {
        "data_granularity": "monthly",
        "monthly_data": monthly_rows,
        "last_full_month_row": dict(ref_row) if ref_row else None,
        "kpi_period": {
            "type": "last_full_month",
            "year": ref_y,
            "month": ref_m,
            "month_name": MONTH_NAMES[ref_m],
        },
        "ytd": {
            "total_plan": ref_row.get("plan") if ref_row else None,
            "total_fact": ref_row.get("fact") if ref_row else None,
            "kpi_pct": ref_row.get("kpi_pct") if ref_row else None,
            "months_with_data": sum(1 for row in monthly_rows if row.get("has_data")),
            "months_total": len(monthly_rows),
            "values_unit": "%",
        },
        "debug": {
            "kpi_id": "HRD-M4",
            "status": "ok",
            "rule": "fact row 10, plan row 11; month columns H, K, N...; Excel percent values converted to percent points",
            **debug,
        },
    }


def cache_file_path_for_period(year: int | None = None, month: int | None = None) -> Path:
    return ytd_json_cache.public_cache_path(CACHE_PREFIX, year, month)


def _load_cache(path: Path, *, source_mtime_ns: int, perpetual: bool) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        with path.open("r", encoding="utf-8") as f:
            raw = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    if raw.get("cache_source") != CACHE_SOURCE_TAG:
        return None
    if raw.get("cache_version") != CACHE_VERSION:
        return None
    if raw.get("source_mtime_ns") != source_mtime_ns:
        return None
    payload = raw.get("payload")
    if not isinstance(payload, dict):
        return None
    if perpetual or raw.get("cache_date") == date.today().isoformat():
        return payload
    return None


def _save_cache(path: Path, payload: dict[str, Any], *, source_mtime_ns: int) -> None:
    ytd_json_cache.CACHE_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", encoding="utf-8") as f:
            json.dump(
                {
                    "cache_source": CACHE_SOURCE_TAG,
                    "cache_version": CACHE_VERSION,
                    "cache_date": date.today().isoformat(),
                    "source_mtime_ns": source_mtime_ns,
                    "payload": payload,
                },
                f,
                ensure_ascii=False,
                indent=2,
            )
    except OSError:
        logger.exception("HRD-M4: не удалось сохранить кэш")


def get_hrd_m4_ytd(year: int | None = None, month: int | None = None) -> dict[str, Any] | None:
    ref_y, ref_m = normalize_rd_tile_period(year, month)
    cache_path = ytd_json_cache.cache_path(CACHE_PREFIX, ref_y, ref_m)
    perpetual = ytd_json_cache.is_ref_period_fully_past(ref_y, ref_m)

    def _runner() -> dict[str, Any] | None:
        try:
            source_mtime_ns = SOURCE_FILE.stat().st_mtime_ns
        except OSError:
            logger.exception("HRD-M4: не найден источник %s", SOURCE_FILE)
            return None

        cached = _load_cache(cache_path, source_mtime_ns=source_mtime_ns, perpetual=perpetual)
        if cached is not None:
            return cached

        try:
            payload = _build_payload(year=ref_y, month=ref_m)
        except Exception:
            logger.exception("HRD-M4: ошибка расчёта текучести")
            return None
        _save_cache(cache_path, payload, source_mtime_ns=source_mtime_ns)
        return payload

    return locked_call(f"sup_hrd_m4_{ref_y}_{ref_m:02d}", _runner)
