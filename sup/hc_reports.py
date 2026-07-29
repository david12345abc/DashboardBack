"""Общие пути и чтение HR-отчётов ``HC_сводный_{year}_{Month}.xls[x]``."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import xlrd

HC_REPORTS_DIR = Path(
    r"\\192.168.1.198\Files\16.Отдел персонала\Отдел\Отчеты",
)

HC_FILE_MONTH_TITLES: dict[int, str] = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

# С июля 2026 HR кладёт .xlsx; раньше — .xls. Берём существующий, при обоих — новее.
_HC_SUFFIXES = (".xlsx", ".xls")


def hc_report_path(year: int, month: int) -> Path:
    title = HC_FILE_MONTH_TITLES[month]
    stem = f"HC_сводный_{year}_{title}"
    existing: list[Path] = []
    for suffix in _HC_SUFFIXES:
        path = HC_REPORTS_DIR / f"{stem}{suffix}"
        try:
            if path.exists():
                existing.append(path)
        except OSError:
            continue
    if not existing:
        return HC_REPORTS_DIR / f"{stem}.xls"
    if len(existing) == 1:
        return existing[0]
    return max(existing, key=lambda p: p.stat().st_mtime_ns)


def reports_mtime_ns(ref_y: int, ref_m: int) -> int:
    latest = 0
    for month in range(1, ref_m + 1):
        path = hc_report_path(ref_y, month)
        try:
            if path.exists():
                latest = max(latest, path.stat().st_mtime_ns)
        except OSError:
            continue
    return latest


class _XlsxSheet:
    """Минимальный адаптер листа openpyxl → API как у xlrd.sheet."""

    def __init__(self, rows: list[tuple[Any, ...]]):
        self._rows = rows
        self.nrows = len(rows)

    def row_values(self, rowx: int) -> list[Any]:
        if rowx < 0 or rowx >= self.nrows:
            return []
        return list(self._rows[rowx])

    def cell_value(self, rowx: int, colx: int) -> Any:
        if rowx < 0 or rowx >= self.nrows:
            return ""
        row = self._rows[rowx]
        if colx < 0 or colx >= len(row):
            return ""
        value = row[colx]
        return "" if value is None else value


class _XlsxBook:
    """Минимальный адаптер openpyxl → API как у xlrd.Book."""

    def __init__(self, path: Path):
        from openpyxl import load_workbook

        self._wb = load_workbook(str(path), data_only=True, read_only=True)
        self.datemode = 0
        self._sheets: dict[str, _XlsxSheet] = {}

    def sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    def sheet_by_name(self, name: str) -> _XlsxSheet:
        if name not in self._sheets:
            ws = self._wb[name]
            rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
            self._sheets[name] = _XlsxSheet(rows)
        return self._sheets[name]

    def close(self) -> None:
        try:
            self._wb.close()
        except Exception:
            pass


def open_hc_workbook(path: Path):
    """Открыть HC-отчёт: .xls через xlrd, .xlsx через openpyxl."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _XlsxBook(path)
    return xlrd.open_workbook(str(path))
