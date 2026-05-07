"""Read metrics from an external defect report.

Install dependencies:
    python -m pip install -r requirements.txt

Examples:
    python external_defect_report_total.py --year 2026 --month апрель
    python external_defect_report_total.py --year 2026 --month 04
"""

from __future__ import annotations

import argparse
import logging
import shutil
import tempfile
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


logger = logging.getLogger(__name__)

BASE_REPORTS_PATH = Path(
    r"\\192.168.1.198\Files\Обмен\Управление несоответствиями\Отчеты"
)
REPORT_NAME_PART = "Отчет по внешнему браку 03-18"
# Основной лист; в части файлов вместо него — «Отчёты по внешнему браку» (ё/е).
TARGET_SHEET_CANDIDATES: tuple[str, ...] = (
    "Анализ год выпуска",
    "Отчёты по внешнему браку",
    "Отчеты по внешнему браку",
)
TARGET_CELL_TEXT = "Общий итог"
CLASSIFIER_HEADER_TEXT = "Классификатор брака"
# Наименования типов брака — как на листе «классификатор»; лист из Excel не читаем.
CLASSIFIER_DISPLAY_NAMES: dict[int, str] = {
    1: "Выход из строя электронных компонентов (внешний брак)",
    2: "Несоответствие линейных размеров (Внешнее несоответствие)",
    3: "Несоответствие МХ (внешнее несоответствие)",
    4: (
        "Изготовление прибора, несоответствующее требованиям заказчика "
        "(внешнее несоответствие)"
    ),
    5: "Негерметичность",
    6: "Низкое качество разработки",
    7: "Несоответствие покрытия (Внешнее несоответствие)",
    8: "Некомплектность поставки (комплектующие)",
    9: "Некомплектность поставки (документы)",
    10: "Несоответствие документации (Внешнее несоответствие)",
    11: "Несоответствие в ПО (внешний брак)",
    12: "Мех. повреждения",
    13: "Брак мех.участка",
    14: "Не гарантия",
    15: "Не соответствие покупных ТМЦ",
}
CLASSIFIER_TYPES: tuple[int, ...] = tuple(sorted(CLASSIFIER_DISPLAY_NAMES))
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

MONTH_ALIASES = {
    "1": "январь",
    "01": "январь",
    "январь": "январь",
    "января": "январь",
    "2": "февраль",
    "02": "февраль",
    "февраль": "февраль",
    "февраля": "февраль",
    "3": "март",
    "03": "март",
    "март": "март",
    "марта": "март",
    "4": "апрель",
    "04": "апрель",
    "апрель": "апрель",
    "апреля": "апрель",
    "5": "май",
    "05": "май",
    "май": "май",
    "мая": "май",
    "6": "июнь",
    "06": "июнь",
    "июнь": "июнь",
    "июня": "июнь",
    "7": "июль",
    "07": "июль",
    "июль": "июль",
    "июля": "июль",
    "8": "август",
    "08": "август",
    "август": "август",
    "августа": "август",
    "9": "сентябрь",
    "09": "сентябрь",
    "сентябрь": "сентябрь",
    "сентября": "сентябрь",
    "10": "октябрь",
    "октябрь": "октябрь",
    "октября": "октябрь",
    "11": "ноябрь",
    "ноябрь": "ноябрь",
    "ноября": "ноябрь",
    "12": "декабрь",
    "декабрь": "декабрь",
    "декабря": "декабрь",
}

MONTH_NUMBERS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def normalize_month(month: str) -> str:
    normalized = month.strip().lower().rstrip(".")
    if normalized not in MONTH_ALIASES:
        raise ValueError("Unknown month. Use 1-12 or a Russian month name.")

    return MONTH_ALIASES[normalized]


def normalize_folder_name(name: str) -> str:
    normalized_name = name.strip().casefold().replace("ё", "е")
    return " ".join(normalized_name.replace("_", " ").replace("-", " ").split())


def find_year_folder(base_path: Path, year: int) -> Path:
    year_path = base_path / str(year)
    if year_path.is_dir():
        return year_path

    matches = [path for path in base_path.iterdir() if path.is_dir() and path.name.strip() == str(year)]
    if not matches:
        raise FileNotFoundError(f"Year folder was not found: {year} in {base_path}")

    return matches[0]


def find_month_folder(year_path: Path, month: str) -> Path:
    normalized_month = normalize_month(month)
    month_number = MONTH_NUMBERS[normalized_month]
    month_number_variants = {str(month_number), f"{month_number:02d}"}

    candidates: list[Path] = []
    for path in year_path.iterdir():
        if not path.is_dir():
            continue

        folder_name = normalize_folder_name(path.name)
        folder_parts = set(folder_name.split())
        if normalized_month in folder_name or folder_parts.intersection(month_number_variants):
            candidates.append(path)

    if not candidates:
        raise FileNotFoundError(f"Month folder was not found: {month} in {year_path}")

    candidates.sort(key=lambda path: path.name)
    return candidates[0]


def find_report_file(month_path: Path) -> Path:
    report_name_part = REPORT_NAME_PART.casefold().replace("ё", "е")
    matches = [
        path
        for path in month_path.iterdir()
        if path.is_file()
        and report_name_part in path.name.casefold().replace("ё", "е")
        and path.suffix.lower() in SUPPORTED_EXCEL_SUFFIXES
    ]

    if not matches:
        raise FileNotFoundError(
            f"Excel report containing '{REPORT_NAME_PART}' was not found in {month_path}"
        )

    matches.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return matches[0]


@contextmanager
def local_file_copy(source_path: Path) -> Iterator[Path]:
    if not source_path.exists():
        raise FileNotFoundError(f"Source file was not found: {source_path}")

    with tempfile.TemporaryDirectory(prefix="external_defect_report_") as temp_dir:
        temp_path = Path(temp_dir) / f"report_copy{source_path.suffix}"
        shutil.copy2(source_path, temp_path)
        yield temp_path


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def normalize_text(value: str) -> str:
    return normalize_cell_value(value).casefold().replace("ё", "е")


def parse_number(value: Any) -> float:
    if value is None:
        return 0

    if isinstance(value, int | float):
        return float(value)

    normalized_value = str(value).strip().replace(",", ".")
    if not normalized_value:
        return 0

    try:
        return float(normalized_value)
    except ValueError:
        return 0


def format_number(value: float) -> str:
    if value.is_integer():
        return str(int(value))

    return str(value)


def is_filled_summary_cell(cell: Any) -> bool:
    """
    Подсвеченные сводные ячейки отчёта: ``solid``, либо заливка с заданным ``fgColor``.
    """
    try:
        fill = cell.fill
        if fill is None:
            return False
        if getattr(fill, "fill_type", None) == "solid":
            return True
        fg = getattr(fill, "fgColor", None)
        if fg is None:
            return False
        if getattr(fg, "rgb", None):
            return True
        t = getattr(fg, "type", None)
        return bool(t and str(t).lower() != "none")
    except (AttributeError, TypeError):
        return False


def find_total_value(sheet: Worksheet) -> tuple[str, str]:
    want = normalize_text(TARGET_CELL_TEXT)
    for row in sheet.iter_rows():
        for cell in row:
            if normalize_text(normalize_cell_value(cell.value)) != want:
                continue

            right_cell = sheet.cell(row=cell.row, column=cell.column + 1)
            right_value = normalize_cell_value(right_cell.value)
            if right_value:
                return cell.coordinate, right_value

    raise ValueError(
        f"Cell '{TARGET_CELL_TEXT}' with a non-empty cell to the right was not found "
        f"on sheet '{sheet.title}'"
    )


def find_cell_by_text(sheet: Worksheet, text: str) -> Any:
    normalized_text = normalize_text(text)
    for row in sheet.iter_rows():
        for cell in row:
            if normalize_text(normalize_cell_value(cell.value)) == normalized_text:
                return cell

    raise ValueError(f"Cell '{text}' was not found on sheet '{sheet.title}'")


def find_classifier_section_header_cell(sheet: Worksheet) -> Any:
    """Строка-якорь блока классификаторов (в отчётах текст может немного отличаться)."""
    aliases = (
        CLASSIFIER_HEADER_TEXT,
        "Классификатор",
    )
    seen: set[str] = set()
    for raw in aliases:
        key = normalize_text(raw)
        if key in seen:
            continue
        seen.add(key)
        try:
            return find_cell_by_text(sheet, raw)
        except ValueError:
            continue
    raise ValueError(
        f"No classifier header found (tried: {aliases}) on sheet '{sheet.title}'"
    )


def find_classifier_columns(sheet: Worksheet) -> dict[int, int]:
    header_cell = find_classifier_section_header_cell(sheet)
    header_row = header_cell.row + 1
    columns: dict[int, int] = {}

    for column in range(header_cell.column, sheet.max_column + 1):
        value = normalize_cell_value(sheet.cell(row=header_row, column=column).value)
        if not value:
            continue

        try:
            classifier_type = int(float(value.replace(",", ".")))
        except ValueError:
            continue

        if classifier_type in CLASSIFIER_TYPES:
            columns[classifier_type] = column

    if not columns:
        raise ValueError(
            f"Classifier columns were not found (expected headers {list(CLASSIFIER_TYPES)} "
            f"below «{CLASSIFIER_HEADER_TEXT}»)"
        )

    return columns


def calculate_classifier_totals(sheet: Worksheet) -> dict[int, float]:
    classifier_columns = find_classifier_columns(sheet)
    first_data_row = find_classifier_section_header_cell(sheet).row + 2
    totals = {classifier_type: 0.0 for classifier_type in CLASSIFIER_TYPES}

    for row in range(first_data_row, sheet.max_row + 1):
        for classifier_type, column in classifier_columns.items():
            cell = sheet.cell(row=row, column=column)
            if is_filled_summary_cell(cell):
                totals[classifier_type] += parse_number(cell.value)

    return dict(totals)


def _normalize_sheet_title(title: str) -> str:
    return title.strip().casefold().replace("ё", "е")


def resolve_target_sheet_name(workbook: Any) -> str:
    """
    Имя листа с блоком «Общий итог» и классификатором: по порядку из
    ``TARGET_SHEET_CANDIDATES``, с запасным сопоставлением без учёта регистра и ё/е.
    """
    names = list(workbook.sheetnames)
    name_set = set(names)
    for candidate in TARGET_SHEET_CANDIDATES:
        if candidate in name_set:
            return candidate
    by_norm: dict[str, str] = {}
    for n in names:
        key = _normalize_sheet_title(n)
        if key not in by_norm:
            by_norm[key] = n
    for candidate in TARGET_SHEET_CANDIDATES:
        key = _normalize_sheet_title(candidate)
        if key in by_norm:
            return by_norm[key]
    raise ValueError(
        "Ни один из листов не найден: "
        + ", ".join(repr(s) for s in TARGET_SHEET_CANDIDATES)
        + f". Доступные листы: {', '.join(names)}"
    )


def read_report_metrics(
    report_path: Path,
) -> tuple[str, str, dict[int, float], dict[int, str], str | None]:
    """Последнее значение tuple — предупреждение, если суммы по типам недоступны (итог всё равно есть)."""
    classifier_names = dict(CLASSIFIER_DISPLAY_NAMES)
    with local_file_copy(report_path) as copied_report_path:
        workbook = load_workbook(copied_report_path, data_only=True, read_only=False)
        try:
            target_sheet_title = resolve_target_sheet_name(workbook)
            sheet = workbook[target_sheet_title]
            total_cell, total_value = find_total_value(sheet)
            classifier_warn: str | None = None
            try:
                classifier_totals = calculate_classifier_totals(sheet)
            except (ValueError, TypeError, KeyError, AttributeError) as exc:
                classifier_warn = str(exc)
                logger.warning(
                    "Внешний отчёт %s: итог взят, разрез по классификатору не прочитан: %s",
                    report_path.name,
                    exc,
                )
                classifier_totals = {ct: 0.0 for ct in CLASSIFIER_TYPES}
            return (
                total_cell,
                total_value,
                classifier_totals,
                classifier_names,
                classifier_warn,
            )
        finally:
            workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read the external defect report total for a selected year and month."
    )
    parser.add_argument("--year", type=int, required=True, help="Report year, for example: 2026.")
    parser.add_argument(
        "--month",
        required=True,
        help="Report month, for example: апрель, 4, 04.",
    )
    parser.add_argument(
        "--base-path",
        type=Path,
        default=BASE_REPORTS_PATH,
        help=f"Reports root folder. Default: {BASE_REPORTS_PATH}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    year_path = find_year_folder(args.base_path, args.year)
    month_path = find_month_folder(year_path, args.month)
    report_path = find_report_file(month_path)
    _total_cell, total_value, classifier_totals, classifier_names, cls_warn = read_report_metrics(
        report_path
    )

    print(f"Total: {total_value}")
    if cls_warn:
        print(f"Classifier warning: {cls_warn}")
    print("Classifier totals:")
    for classifier_type in CLASSIFIER_TYPES:
        n = classifier_totals[classifier_type]
        print(f"{classifier_names[classifier_type]}: {format_number(n)}")


if __name__ == "__main__":
    main()
