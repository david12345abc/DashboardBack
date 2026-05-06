"""
QD-M1 — уровень внешнего брака / рекламаций (директор по качеству).

Итог за месяц: последняя строка с «всего»/«итого»; план — колонка 3,
факт — колонка 4.

Разрез по статьям (поле ``articles``): строка считается статьёй, если
выполняется любое из правил:

- колонка 1 пустая, в колонке 2 — наименование, колонки 3 и 4 — план и факт;
- **объединённые ячейки** в одной строке: текст названия в области merge,
  в **двух следующих** колонках после правой границы merge — план и факт;
- строка **выделена зелёным** (заливка), при этом наименование и план/факт
  берутся как у «пустой кол.1 + кол.2» или «текст в кол.1», колонки 3–4.

Строка итога («всего»/«итого») и шапки таблицы не включаются.
"""

from __future__ import annotations

import os
import re
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

from getkpi.cache_manager import locked_call
from getkpi.techdir_tekuchet import MONTH_RU

SOURCE_TAG = "qualdir_qd_m1_xlsx_v3"

_TOTAL_LABEL_RE = re.compile(r"\b(?:всего|итого)\b", re.IGNORECASE)
_DATE_IN_TITLE_RE = re.compile(r"(\d{1,2})\.(\d{1,2})\.(\d{4})")
# Строки-шапки таблицы (колонка 1 пустая, в колонке 2 — не статья, а заголовок).
_ARTICLE_SKIP_NAMES = frozenset({
    "наименование",
    "показатель",
    "план",
    "факт",
    "статья",
})


def resolve_qd_m1_xlsx_path() -> Path | None:
    """Путь к файлу: QD_M1_XLSX_PATH или <корень проекта>/QD-M1.xlsx."""
    env = (os.environ.get("QD_M1_XLSX_PATH") or "").strip()
    if env:
        p = Path(env)
        if p.is_file():
            return p
    root = Path(__file__).resolve().parent.parent
    candidate = root / "QD-M1.xlsx"
    if candidate.is_file():
        return candidate
    return None


def _parse_period_year_month_from_title(title: str) -> tuple[int, int] | None:
    """Берёт месяц/год по последней дате вида ДД.ММ.ГГГГ в заголовке листа."""
    matches = list(_DATE_IN_TITLE_RE.finditer(title))
    if not matches:
        return None
    m = matches[-1]
    _d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if not (1 <= mo <= 12):
        return None
    return y, mo


def _cell_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        x = float(value)
        return x if x == x else None
    s = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if s.endswith("%"):
        s = s[:-1].strip()
    if not s:
        return None
    try:
        x = float(s)
        return x if x == x else None
    except ValueError:
        return None


def _row_has_total_label(row: tuple[Any, ...]) -> bool:
    for cell in row:
        if cell is None:
            continue
        s = str(cell).strip()
        if not s:
            continue
        if _TOTAL_LABEL_RE.search(s):
            return True
    return False


def _col1_empty(cell: Any) -> bool:
    """Колонка 1 «пустая»: None, пустая строка, только пробелы."""
    if cell is None:
        return True
    if isinstance(cell, str) and not cell.strip():
        return True
    return False


def _pad_row(row: tuple[Any, ...], width: int = 4) -> tuple[Any, ...]:
    r = tuple(row)
    if len(r) >= width:
        return r[:width]
    return r + (None,) * (width - len(r))


def _is_article_row(row: tuple[Any, ...]) -> bool:
    """Статья: колонка 1 пустая, в колонке 2 — наименование (не шапка)."""
    c1, c2, _c3, _c4 = _pad_row(row, 4)
    if not _col1_empty(c1):
        return False
    if c2 is None:
        return False
    name = str(c2).strip()
    if not name:
        return False
    if name.lower() in _ARTICLE_SKIP_NAMES:
        return False
    return True


def _cell_fg_rgb(cell: Any) -> str | None:
    """ARGB из заливки ячейки, если solid + rgb."""
    try:
        fill = cell.fill
        if fill is None or fill.patternType != "solid":
            return None
        fg = fill.fgColor
        if fg is None or fg.type != "rgb" or not fg.rgb:
            return None
        return str(fg.rgb).upper()
    except (AttributeError, TypeError):
        return None


def _is_green_rgb(rgb: str) -> bool:
    """Распознавание зелёной темы Excel / типичных зелёных ARGB."""
    r = rgb.upper().replace("#", "")
    if len(r) == 8 and r.startswith("FF"):
        r = r[2:]
    if len(r) != 6:
        return False
    # Excel «зелёный» акцент, листовые зелёные
    greens = {
        "92D050", "00B050", "00FF00", "008000", "548235", "375623",
        "70AD47", "A9D08E", "C6E0B4",
    }
    if r in greens:
        return True
    rr, gg, bb = int(r[0:2], 16), int(r[2:4], 16), int(r[4:6], 16)
    return gg >= 140 and gg > rr + 20 and gg > bb + 20


def _row_is_green(ws: Any, row_idx: int, max_col: int = 8) -> bool:
    for c in range(1, max_col + 1):
        rgb = _cell_fg_rgb(ws.cell(row_idx, c))
        if rgb and _is_green_rgb(rgb):
            return True
    return False


def _article_dict(
    name: str,
    plan: Any,
    fact: Any,
    excel_row: int,
) -> dict[str, Any] | None:
    name = name.strip()
    if not name or name.lower() in _ARTICLE_SKIP_NAMES:
        return None
    pv, fv = _cell_float(plan), _cell_float(fact)
    if pv is None and fv is None:
        return None
    return {
        "name": name,
        "plan": pv,
        "fact": fv,
        "excel_row": excel_row,
    }


def _article_from_merged_row(ws: Any, row_idx: int) -> dict[str, Any] | None:
    """Название в объединённых ячейках одной строки; план и факт — две колонки справа от merge."""
    max_col = ws.max_column or 0
    for mr in ws.merged_cells.ranges:
        if mr.min_row != row_idx or mr.max_row != row_idx:
            continue
        plan_col = mr.max_col + 1
        fact_col = mr.max_col + 2
        if plan_col > max_col:
            continue
        name_cell = ws.cell(row_idx, mr.min_col)
        raw_name = name_cell.value
        if raw_name is None:
            continue
        name = str(raw_name).strip()
        plan_v = ws.cell(row_idx, plan_col).value
        fact_v = ws.cell(row_idx, fact_col).value if fact_col <= max_col else None
        art = _article_dict(name, plan_v, fact_v, row_idx)
        if art:
            return art
    return None


def _article_from_legacy_row(row_vals: tuple[Any, ...], row_idx: int) -> dict[str, Any] | None:
    c1, c2, c3, c4 = _pad_row(row_vals, 4)
    if not _is_article_row((c1, c2, c3, c4)):
        return None
    return _article_dict(str(c2).strip(), c3, c4, row_idx)


def _article_from_green_row(
    ws: Any,
    row_idx: int,
    row_vals: tuple[Any, ...],
) -> dict[str, Any] | None:
    if not _row_is_green(ws, row_idx):
        return None
    c1, c2, c3, c4 = _pad_row(row_vals, 4)
    if _col1_empty(c1) and c2 is not None and str(c2).strip():
        return _article_dict(str(c2).strip(), c3, c4, row_idx)
    if c1 is not None and str(c1).strip():
        return _article_dict(str(c1).strip(), c3, c4, row_idx)
    return None


def _row_tuple_values(ws: Any, row_idx: int, max_col: int) -> tuple[Any, ...]:
    return tuple(ws.cell(row_idx, c).value for c in range(1, max_col + 1))


def _parse_workbook(path: Path) -> dict[str, Any]:
    # Нужны формулы как значения, merged_cells и заливка — не read_only.
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb.active
        first_row = _row_tuple_values(ws, 1, min(ws.max_column or 8, 12))
        title = str(first_row[0] if first_row else "") or ""

        period = _parse_period_year_month_from_title(title)
        if period is None:
            today = date.today()
            period = (today.year, 4)

        data_year, data_month = period

        last_plan: float | None = None
        last_fact: float | None = None
        last_row_idx = 0
        articles: list[dict[str, Any]] = []

        max_row = ws.max_row or 0
        scan_cols = min(ws.max_column or 8, 16)

        for idx in range(2, max_row + 1):
            row_vals = _row_tuple_values(ws, idx, scan_cols)
            if _row_has_total_label(row_vals):
                c3 = ws.cell(idx, 3).value
                c4 = ws.cell(idx, 4).value
                last_plan = _cell_float(c3)
                last_fact = _cell_float(c4)
                last_row_idx = idx
                continue

            art = _article_from_merged_row(ws, idx)
            if art is None:
                art = _article_from_legacy_row(row_vals, idx)
            if art is None:
                art = _article_from_green_row(ws, idx, row_vals)
            if art is not None:
                articles.append(art)

        ok = last_row_idx > 0 and (last_plan is not None or last_fact is not None)
        return {
            "ok": ok,
            "data_year": data_year,
            "data_month": data_month,
            "plan": last_plan,
            "fact": last_fact,
            "excel_row": last_row_idx,
            "title_sample": title[:200],
            "source_file": str(path),
            "articles": articles,
        }
    finally:
        wb.close()


def _qd_m1_kpi_pct(plan: Any, fact: Any) -> float | None:
    if plan is None or fact is None:
        return None
    try:
        pv = float(plan)
        fv = float(fact)
    except (TypeError, ValueError):
        return None
    if pv <= 0:
        return None
    return round(fv / pv * 100.0, 1)


def _month_pairs_through(year: int, ref_month: int) -> list[tuple[int, int]]:
    return [(year, mm) for mm in range(1, ref_month + 1)]


def _last_full_month() -> tuple[int, int]:
    today = date.today()
    if today.month == 1:
        return today.year - 1, 12
    return today.year, today.month - 1


def _build_payload(year: int, month: int, parsed: dict[str, Any]) -> dict[str, Any]:
    dy, dm = int(parsed["data_year"]), int(parsed["data_month"])
    plan_src = parsed.get("plan")
    fact_src = parsed.get("fact")
    has_file = parsed.get("ok")
    raw_articles: list[dict[str, Any]] = list(parsed.get("articles") or [])
    # Разрез из того же XLSX, что и итог; отдаём при успешном разборе файла (месяц снимка — data_year/data_month).
    articles_out: list[dict[str, Any]] = []
    if has_file:
        articles_out = [
            {"name": a["name"], "plan": a["plan"], "fact": a["fact"]}
            for a in raw_articles
        ]

    month_rows: list[dict[str, Any]] = []
    for row_y, row_m in _month_pairs_through(year, month):
        if has_file and row_y == dy and row_m == dm:
            plan, fact = plan_src, fact_src
            has_data = plan is not None and fact is not None
            month_rows.append({
                "year": row_y,
                "month": row_m,
                "month_name": MONTH_RU[row_m].lower(),
                "plan": plan,
                "fact": fact,
                "kpi_pct": _qd_m1_kpi_pct(plan, fact),
                "has_data": has_data,
            })
        else:
            month_rows.append({
                "year": row_y,
                "month": row_m,
                "month_name": MONTH_RU[row_m].lower(),
                "plan": None,
                "fact": None,
                "kpi_pct": None,
                "has_data": False,
            })

    ref_row = next(
        (r for r in month_rows if r["year"] == year and r["month"] == month),
        month_rows[-1] if month_rows else None,
    )

    with_data = [r for r in month_rows if r.get("has_data")]
    months_with_data = len(with_data)

    return {
        "data_granularity": "monthly",
        "monthly_data": month_rows,
        "last_full_month_row": dict(ref_row) if ref_row else None,
        "ytd": {
            "total_plan": ref_row.get("plan") if ref_row else None,
            "total_fact": ref_row.get("fact") if ref_row else None,
            "kpi_pct": ref_row.get("kpi_pct") if ref_row else None,
            "months_with_data": months_with_data,
            "months_total": len(month_rows),
        },
        "kpi_period": {
            "type": "last_full_month",
            "year": year,
            "month": month,
            "month_name": MONTH_RU[month],
            "data_complete": bool(ref_row and ref_row.get("has_data")),
        },
        "articles": articles_out,
        "debug": {
            "status": "ok" if has_file else "no_data",
            "kpi_id": "QD-M1",
            "source": SOURCE_TAG,
            "parsed": {k: v for k, v in parsed.items() if k not in {"ok", "articles"}},
        },
    }


def get_qd_m1_ytd(year: int | None = None, month: int | None = None) -> dict[str, Any]:
    """Помесячный payload для плитки QD-M1 (данные из XLSX за месяц из заголовка файла)."""

    def _runner() -> dict[str, Any]:
        ry, rm = year, month
        if ry is None or rm is None:
            ry, rm = _last_full_month()
        try:
            path = resolve_qd_m1_xlsx_path()
            if path is None:
                y, m = ry, rm
                return {
                    "data_granularity": "monthly",
                    "monthly_data": [],
                    "last_full_month_row": None,
                    "ytd": {
                        "total_plan": None,
                        "total_fact": None,
                        "kpi_pct": None,
                        "months_with_data": 0,
                        "months_total": 0,
                    },
                    "kpi_period": {
                        "type": "last_full_month",
                        "year": y,
                        "month": m,
                        "month_name": MONTH_RU[m],
                        "data_complete": False,
                    },
                    "articles": [],
                    "debug": {
                        "status": "error",
                        "kpi_id": "QD-M1",
                        "source": SOURCE_TAG,
                        "error": "Файл QD-M1.xlsx не найден (или задайте QD_M1_XLSX_PATH).",
                    },
                }

            parsed = _parse_workbook(path)
            if not parsed.get("ok"):
                y, m = ry, rm
                return {
                    "data_granularity": "monthly",
                    "monthly_data": [],
                    "last_full_month_row": None,
                    "ytd": {
                        "total_plan": None,
                        "total_fact": None,
                        "kpi_pct": None,
                        "months_with_data": 0,
                        "months_total": 0,
                    },
                    "kpi_period": {
                        "type": "last_full_month",
                        "year": y,
                        "month": m,
                        "month_name": MONTH_RU[m],
                        "data_complete": False,
                    },
                    "articles": [],
                    "debug": {
                        "status": "error",
                        "kpi_id": "QD-M1",
                        "source": SOURCE_TAG,
                        "error": "Не удалось найти строку «всего»/«итого» с планом/фактом в колонках 3–4.",
                        "parsed": parsed,
                    },
                }

            return _build_payload(ry, rm, parsed)
        except Exception as exc:
            y, m = ry, rm
            return {
                "data_granularity": "monthly",
                "monthly_data": [],
                "last_full_month_row": None,
                "ytd": {
                    "total_plan": None,
                    "total_fact": None,
                    "kpi_pct": None,
                    "months_with_data": 0,
                    "months_total": 0,
                },
                "kpi_period": {
                    "type": "last_full_month",
                    "year": y,
                    "month": m,
                    "month_name": MONTH_RU[m],
                    "data_complete": False,
                },
                "articles": [],
                "debug": {
                    "status": "error",
                    "kpi_id": "QD-M1",
                    "source": SOURCE_TAG,
                    "error": str(exc),
                },
            }

    lock_y, lock_m = year, month
    if lock_y is None or lock_m is None:
        lock_y, lock_m = _last_full_month()
    return locked_call(f"qualdir_qd_m1_{lock_y}_{lock_m:02d}", _runner)


def qd_m1_excel_paths_for_cache_stamp() -> list[Path]:
    p = resolve_qd_m1_xlsx_path()
    return [p] if p is not None else []
