# -*- coding: utf-8 -*-
from pathlib import Path
from openpyxl import load_workbook

base = Path(r"C:\Users\testii\Downloads\dash\DashboardBack\temp")
files = sorted(base.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
for p in files:
    if p.name.startswith("dogovory"):
        continue
    print("=" * 80)
    print("FILE", p.name, p.stat().st_size)
    try:
        wb = load_workbook(p, data_only=True, read_only=True)
    except Exception as e:
        print(" open err", e)
        continue
    print(" sheets", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        # print first 40 rows looking for договор/ожид/план
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            vals = [c for c in row if c is not None]
            if not vals:
                continue
            text = " | ".join(str(v) for v in vals[:12])
            low = text.lower()
            if i <= 25 or any(
                k in low
                for k in (
                    "договор",
                    "ожид",
                    "маркетинг",
                    "итог",
                    "отдел",
                    "факт",
                    "план",
                )
            ):
                rows.append((i, text[:240]))
            if i > 120:
                break
        print(" --", sn)
        for i, t in rows[:60]:
            print(f"  {i}: {t}")
    wb.close()
